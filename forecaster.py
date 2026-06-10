"""
Forecaster Module - High-level forecasting interface with progress tracking
"""

import pandas as pd
import numpy as np
from typing import Optional, List, Dict, Tuple, Any
import logging
from pathlib import Path

# Handle both relative and absolute imports
try:
    from .models import get_available_models
except ImportError:
    from models import get_available_models

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Tuning constants  (documented rationale for each threshold)
# ---------------------------------------------------------------------------

# Soft floor for data-sufficiency checks: allows forecasting sparse site/grade
# combos with as few as 6 months while recommending 24.  Clamped to [6, 12].
SOFT_MIN_MONTHS_FLOOR = 6
SOFT_MIN_MONTHS_CEIL = 12

# Rolling-window size (months) for outlier statistics.  18 months captures one
# full seasonal cycle plus 6 months of trend, keeping the detector responsive
# to legitimate business changes while smoothing noise.
OUTLIER_WINDOW_MONTHS = 18

# Rolling-median window used inside the spike detector.
SPIKE_ROLLING_WINDOW = 6

# Spike multiplier: a single month exceeding 8x the rolling median is almost
# certainly a data error (double-scanned load, unit mismatch, etc.).  A lower
# cap (e.g. 5x) clipped legitimate seasonal peaks in backtests.
SPIKE_MULTIPLIER = 8

# Minimum consecutive trailing months above a threshold before we treat the
# pattern as a real business level-shift rather than data errors.
LEVEL_SHIFT_MIN_TRAILING = 3

# Asymmetric MAD bounds for outlier detection.  The lower bound is tighter
# (4x MAD) because near-zero months are almost always errors (system outages,
# store closures misrecorded as zero sales).  The upper bound is looser
# (6x MAD) because legitimate upside surprises (new customer, promo) are more
# common than downside errors.
MAD_LOWER_MULTIPLIER = 4.0
MAD_UPPER_MULTIPLIER = 6.0

# Degenerate-series fallback bands when MAD == 0 (constant series).
DEGENERATE_LOWER_FACTOR = 0.5
DEGENERATE_UPPER_FACTOR = 2.0


# ---------------------------------------------------------------------------
# Post-forecast sanity caps & YoY guardrails
# ---------------------------------------------------------------------------

# Sites with fewer than this many months of data use grade-cohort bounds
# instead of their own history for sanity caps.
MATURE_SITE_MONTHS = 12

# Sanity cap multipliers applied to the reference max/min volume.
SANITY_CAP_UPPER_MULT = 3.0
SANITY_CAP_LOWER_MULT = 0.1

# When extrapolating trend for new sites, the extrapolated value itself
# is capped at this multiple of the most recent actual volume before the
# SANITY_CAP_UPPER_MULT multiplier is applied.
TREND_EXTRAP_CEILING_MULT = 2.0

# Year-over-year guardrails: forecast must stay within these ratios of
# the prior-year same-month actual.
YOY_GUARD_UPPER_RATIO = 3.0
YOY_GUARD_LOWER_RATIO = 0.33


def _normalize_filter(value: Optional[str]) -> Optional[str]:
    """Coerce empty/whitespace-only strings to None for site_id / grade filters."""
    if isinstance(value, str) and value.strip() == "":
        return None
    return value


def _reorder_columns(df: pd.DataFrame, desired_order: List[str]) -> pd.DataFrame:
    """Reorder *df* columns: *desired_order* first, then any remaining."""
    ordered = [c for c in desired_order if c in df.columns]
    ordered_set = set(ordered)
    remaining = [c for c in df.columns if c not in ordered_set]
    return df[ordered + remaining]


class FuelForecaster:
    """Main forecasting interface"""

    # Favor ETS in the default ensemble because it has shown stronger
    # average performance and lower under-forecast bias in backtests.
    DEFAULT_ENSEMBLE_WEIGHTS = {"ets": 0.7, "snaive": 0.3}

    # Shared column schemas for export (used by both Excel and CSV exporters)
    CLEAN_COLS = (
        "site_id",
        "grade",
        "target_month",
        "forecast_volume",
        "forecast_volume_unclamped",
        "sanity_cap_applied",
        "upper_cap_value",
        "lower_floor_value",
        "yoy_cap_applied",
        "forecast_p10",
        "forecast_p90",
        "interval_width_pct",
        "risk_flag",
        "prior_year_volume",
        "yoy_change_pct",
    )
    DETAIL_COLS = (
        "site_id",
        "grade",
        "target_month",
        "model",
        "forecast_volume",
        "forecast_volume_unclamped",
        "sanity_cap_applied",
        "upper_cap_value",
        "lower_floor_value",
        "yoy_cap_applied",
        "forecast_p10",
        "forecast_p90",
        "interval_width_pct",
        "risk_flag",
        "prior_year_month",
        "prior_year_volume",
        "yoy_change_pct",
    )

    # Risk flag threshold: intervals wider than this % are flagged
    RISK_THRESHOLD_PCT = 30.0

    def __init__(self, database, min_months_data: int = 24,
                 use_adaptive_weights: bool = True):
        """
        Initialize forecaster

        Args:
            database: FuelDatabase instance
            min_months_data: Minimum months of data required (default: 24)
            use_adaptive_weights: Use calibrated per-site weights if available (default: True)
        """
        self.db = database
        self.min_months_data = min_months_data
        self.models = {}
        self.ensemble_weights = self.DEFAULT_ENSEMBLE_WEIGHTS.copy()
        self.use_adaptive_weights = use_adaptive_weights
        self._site_weights_cache = None
        self._interval_cache = {}
        # Softer floor for sparse site/grade combos; allows forecasting with fewer months
        self.soft_min_months = max(
            SOFT_MIN_MONTHS_FLOOR, min(SOFT_MIN_MONTHS_CEIL, self.min_months_data)
        )

    def _load_site_weights(self) -> Dict[str, Dict[str, float]]:
        """Lazy-load calibrated site weights from the database."""
        if self._site_weights_cache is None:
            if self.db is not None and self.use_adaptive_weights:
                try:
                    self._site_weights_cache = self.db.get_site_weights_bulk()
                except Exception:
                    self._site_weights_cache = {}
            else:
                self._site_weights_cache = {}
        return self._site_weights_cache

    def _get_interval_factors(self, segment: str) -> Optional[Dict]:
        """Lookup interval factors with caching."""
        if not self.use_adaptive_weights or self.db is None:
            return None
        if segment in self._interval_cache:
            return self._interval_cache[segment]
        try:
            factors = self.db.get_interval_factors(segment)
        except Exception:
            factors = None
        self._interval_cache[segment] = factors
        return factors

    def _compute_intervals(
        self,
        forecast: float,
        site_id: Optional[str] = None,
        grade: Optional[str] = None,
        risk_threshold: float = None,
    ) -> Dict[str, Any]:
        """
        Compute prediction intervals from calibrated residual distributions.

        Shrinkage fallback: site -> grade -> global.
        Returns dict with forecast_p10, forecast_p90, interval_width_pct, risk_flag.
        Returns empty values if no calibration data exists.
        """
        if risk_threshold is None:
            risk_threshold = self.RISK_THRESHOLD_PCT

        factors = None
        # Try site-level first
        if site_id:
            factors = self._get_interval_factors(f"site:{site_id}")
        # Try grade-level fallback if that segment exists
        if factors is None and grade and grade != "ALL":
            factors = self._get_interval_factors(f"grade:{grade}")
        # Try global
        if factors is None:
            factors = self._get_interval_factors("global")

        if factors is None:
            return {
                "forecast_p10": None,
                "forecast_p90": None,
                "interval_width_pct": None,
                "risk_flag": None,
            }

        p10 = max(0.0, forecast * (1 + factors["residual_p10"]))
        p90 = forecast * (1 + factors["residual_p90"])
        width_pct = (p90 - p10) / forecast * 100 if forecast > 0 else None
        risk_flag = "HIGH" if width_pct is not None and width_pct > risk_threshold else ""

        return {
            "forecast_p10": round(p10, 1),
            "forecast_p90": round(p90, 1),
            "interval_width_pct": round(width_pct, 1) if width_pct is not None else None,
            "risk_flag": risk_flag,
        }

    def _compute_ensemble_forecast(
        self, model_forecasts: Dict[str, float], site_id: Optional[str] = None
    ) -> float:
        """Compute a weighted ensemble with adaptive weights and robust fallback."""
        if not model_forecasts:
            raise ValueError("Cannot compute ensemble without model forecasts")

        # Try site-specific calibrated weights
        weights = self.ensemble_weights
        if site_id and self.use_adaptive_weights:
            site_weights = self._load_site_weights()
            if site_id in site_weights:
                weights = site_weights[site_id]

        weighted_sum = 0.0
        weight_total = 0.0
        for model_name, value in model_forecasts.items():
            weight = weights.get(model_name, 0.0)
            if weight > 0:
                weighted_sum += weight * value
                weight_total += weight

        if weight_total > 0:
            return float(weighted_sum / weight_total)

        # Fallback: preserve robust behavior for unexpected model sets.
        return float(np.median(list(model_forecasts.values())))

    def prepare_monthly_data(
        self,
        site_id: Optional[str] = None,
        grade: Optional[str] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        handle_outliers: bool = True,
        fill_gaps: bool = True,
    ) -> pd.DataFrame:
        """
        Prepare monthly aggregated data with outlier detection

        Args:
            site_id: Optional site ID filter
            grade: Optional grade filter
            start_date: Optional start date
            end_date: Optional end date
            handle_outliers: Apply outlier detection/handling (default: True)
            fill_gaps: Fill missing months via interpolation (default: True).
                       Set False for snaive to preserve exact historical values.

        Returns:
            DataFrame with monthly volume data
        """
        site_id = _normalize_filter(site_id)
        grade = _normalize_filter(grade)

        site_ids = [site_id] if site_id else None
        grades = [grade] if grade else None

        df = self.db.get_sales_data(
            start_date=start_date,
            end_date=end_date,
            site_ids=site_ids,
            grades=grades,
            exclude_estimated=True,
        )

        if df.empty:
            raise ValueError("No data available")

        # Aggregate to monthly
        df["date"] = pd.to_datetime(df["day"])
        df["year_month"] = df["date"].dt.to_period("M")

        monthly = (
            df.groupby("year_month")
            .agg({"volume": "sum"})
            .reset_index()
        )

        monthly["date"] = monthly["year_month"].dt.to_timestamp()
        monthly = monthly.sort_values("date").reset_index(drop=True)

        result = monthly[["date", "volume"]].copy()

        # Fill missing months to keep seasonality aligned on sparse series
        # Skip for snaive which needs exact historical values without interpolation
        if fill_gaps:
            result = self._fill_monthly_gaps(result, site_id=site_id, grade=grade)

        # Outlier detection and handling (critical for individual site accuracy)
        if handle_outliers and len(result) >= 12:
            result = self._handle_outliers(result, site_id)

        return result

    def _fill_monthly_gaps(
        self, data: pd.DataFrame, site_id: Optional[str] = None, grade: Optional[str] = None
    ) -> pd.DataFrame:
        """Reindex to full monthly range and fill missing months."""
        if data.empty:
            return data

        start = data["date"].min()
        end = data["date"].max()
        full_range = pd.date_range(start=start, end=end, freq="MS")

        reindexed = data.set_index("date").reindex(full_range)
        missing = int(reindexed["volume"].isna().sum())

        if missing > 0:
            reindexed["volume"] = (
                reindexed["volume"]
                .interpolate(limit_direction="both")
                .ffill()
                .bfill()
            )
            reindexed["volume"] = reindexed["volume"].clip(lower=0.0)

            site_label = f"{site_id}" if site_id else "ALL"
            grade_label = f"{grade}" if grade else "ALL"
            logger.info(
                f"Filled {missing} missing month(s) for site {site_label}, grade {grade_label}"
            )

        reindexed = reindexed.reset_index().rename(columns={"index": "date"})
        return reindexed

    @staticmethod
    def _count_trailing_true(mask) -> int:
        """Count consecutive True values from the end of a boolean array/Series."""
        count = 0
        values = mask if isinstance(mask, np.ndarray) else mask.values
        for val in reversed(values):
            if val:
                count += 1
            else:
                break
        return count

    def _handle_outliers(
        self, data: pd.DataFrame, site_id: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Detect and cap outliers using a single rolling-window MAD method.

        Robust (median/MAD) bounds are computed over the trailing
        OUTLIER_WINDOW_MONTHS, which adapts to legitimate business changes
        (growth, new customers) while catching true data errors (typos, glitches).
        The upper bound additionally respects a per-month spike ceiling
        (SPIKE_MULTIPLIER x local rolling median) so a one-month spike is caught
        even when the window MAD is wide.

        Sustained level shifts (LEVEL_SHIFT_MIN_TRAILING+ consecutive trailing
        months outside the bounds) are preserved, as they represent real business
        changes like high-speed diesel upgrades or new store ramp-ups.

        Outliers are capped at the boundary rather than removed to preserve
        time-series continuity.
        """
        df = data.copy()
        site_label = f"Site {site_id}" if site_id else "Dataset"

        if len(df) < 4:
            return df

        # Robust MAD bounds over the trailing window.
        window_size = min(OUTLIER_WINDOW_MONTHS, len(df))
        recent_volumes = df.tail(window_size)["volume"].values
        median = np.median(recent_volumes)
        MAD = np.median(np.abs(recent_volumes - median))

        if not (MAD > 0 and np.isfinite(MAD)):
            lower_bound = max(0.0, median * DEGENERATE_LOWER_FACTOR)
            upper_bound = median * DEGENERATE_UPPER_FACTOR
        else:
            lower_bound = max(0.0, median - MAD_LOWER_MULTIPLIER * MAD)
            upper_bound = median + MAD_UPPER_MULTIPLIER * MAD

        # Per-month spike ceiling relative to the local rolling median: catches a
        # single extreme month even when the window MAD bound is wide.  The
        # effective upper bound per month is the tighter of the two.
        rolling_median = (
            df["volume"]
            .rolling(window=SPIKE_ROLLING_WINDOW, min_periods=1)
            .median()
            .replace(0, np.nan)
        )
        spike_ceiling = (rolling_median * SPIKE_MULTIPLIER).fillna(
            df["volume"].max() * 2
        )
        effective_upper = np.minimum(upper_bound, spike_ceiling.values)

        volumes = df["volume"].values
        outlier_mask = (volumes < lower_bound) | (volumes > effective_upper)

        trailing = self._count_trailing_true(outlier_mask)
        if trailing >= LEVEL_SHIFT_MIN_TRAILING:
            outlier_mask[-trailing:] = False
            logger.info(
                f"  {site_label}: Step-change detected - preserving "
                f"trailing {trailing} month(s) from outlier bounds "
                f"[{lower_bound:.0f}, {upper_bound:.0f}]"
            )

        outlier_indices = df.index[outlier_mask]
        outlier_count = len(outlier_indices)

        if outlier_count > 0:
            df.loc[outlier_indices, "volume"] = np.clip(
                volumes[outlier_mask], lower_bound, effective_upper[outlier_mask]
            )

            outlier_dates = (
                df.loc[outlier_indices, "date"].dt.strftime("%Y-%m").tolist()
            )
            logger.info(
                f"  {site_label}: Capped {outlier_count} outlier(s) at boundaries "
                f"[{lower_bound:.0f}, {upper_bound:.0f}] in months: {outlier_dates[:5]}"
                f"{'...' if len(outlier_dates) > 5 else ''}"
            )

        return df

    def _assess_data_quality(self, months_available: int) -> Tuple[str, Optional[str]]:
        """
        Categorize data sufficiency for transparency in outputs.
        Returns (label, note).
        """
        if months_available >= self.min_months_data:
            return "ok", None
        if months_available >= self.soft_min_months:
            return (
                "sparse",
                f"{months_available} months (<{self.min_months_data})",
            )
        return (
            "very_sparse",
            f"{months_available} months (<{self.min_months_data}); high uncertainty",
        )

    def _get_prior_year_actual(
        self,
        target_month: str,
        site_id: Optional[str] = None,
        grade: Optional[str] = None,
    ) -> Optional[float]:
        """
        Get the actual volume for the same month in the prior year.

        Args:
            target_month: Target month in 'YYYY-MM' format
            site_id: Optional site ID filter
            grade: Optional grade filter

        Returns:
            Actual volume for prior year same month, or None if not available
        """
        target_date = pd.to_datetime(target_month)
        prior_year_month = (target_date - pd.DateOffset(years=1)).strftime("%Y-%m")
        prior_year_start = f"{prior_year_month}-01"
        prior_year_end = (
            pd.to_datetime(prior_year_start) + pd.offsets.MonthEnd(0)
        ).strftime("%Y-%m-%d")

        site_ids = [site_id] if site_id else None
        grades = [grade] if grade else None

        try:
            df = self.db.get_sales_data(
                start_date=prior_year_start,
                end_date=prior_year_end,
                site_ids=site_ids,
                grades=grades,
                exclude_estimated=True,
            )

            if df.empty:
                return None

            return float(df["volume"].sum())
        except Exception as e:
            logger.debug(f"Could not fetch prior year actual: {e}")
            return None

    def _precompute_prior_year_actuals(
        self, target_month: str
    ) -> Dict[Tuple[str, str], float]:
        """
        Fetch prior-year actuals for ALL site/grade combos in a single query.

        Returns:
            Dict keyed by (site_id, grade) -> volume.
            site_id or grade may be "ALL" for aggregate-level lookups.
        """
        target_date = pd.to_datetime(target_month)
        prior_year_month = (target_date - pd.DateOffset(years=1)).strftime("%Y-%m")
        prior_year_start = f"{prior_year_month}-01"
        prior_year_end = (
            pd.to_datetime(prior_year_start) + pd.offsets.MonthEnd(0)
        ).strftime("%Y-%m-%d")

        try:
            df = self.db.get_sales_data(
                start_date=prior_year_start,
                end_date=prior_year_end,
                exclude_estimated=True,
            )
        except Exception as e:
            logger.debug(f"Could not fetch prior year actuals in bulk: {e}")
            return {}

        if df.empty:
            return {}

        lookup: Dict[Tuple[str, str], float] = {}

        # Per site+grade
        by_sg = df.groupby(["site_id", "grade"])["volume"].sum()
        for (sid, gr), vol in by_sg.items():
            lookup[(str(sid), str(gr))] = float(vol)

        # Per site (grade=ALL)
        by_site = df.groupby("site_id")["volume"].sum()
        for sid, vol in by_site.items():
            lookup[(str(sid), "ALL")] = float(vol)

        # Per grade (site=ALL)
        by_grade = df.groupby("grade")["volume"].sum()
        for gr, vol in by_grade.items():
            lookup[("ALL", str(gr))] = float(vol)

        # Grand total
        lookup[("ALL", "ALL")] = float(df["volume"].sum())

        return lookup

    def _calculate_yoy_change(
        self, forecast_volume: float, prior_year_volume: Optional[float]
    ) -> Optional[float]:
        """
        Calculate year-over-year percentage change.

        Args:
            forecast_volume: Forecasted volume for target month
            prior_year_volume: Actual volume from same month prior year

        Returns:
            Percentage change (e.g., 5.2 for 5.2% increase), or None if prior year unavailable
        """
        # Handle None, NaN, and zero values to avoid division errors
        if prior_year_volume is None or pd.isna(prior_year_volume) or prior_year_volume == 0:
            return None
        return ((forecast_volume - prior_year_volume) / prior_year_volume) * 100

    def _get_prior_year_actual_by_grade(
        self, target_month: str, grade: str
    ) -> Optional[float]:
        """
        Get actual aggregate volume for a specific grade across all sites for prior year month.

        This queries the database directly for the true historical aggregate,
        ensuring accurate YoY comparisons in summary reports.

        Args:
            target_month: Target month in 'YYYY-MM' format
            grade: Fuel grade to filter by

        Returns:
            Actual total volume for the grade in prior year same month, or None if unavailable
        """
        # Handle "ALL" grade (used in site-level forecasts) by fetching total across all grades
        if grade == "ALL":
            return self._get_prior_year_actual_total(target_month)
        return self._get_prior_year_actual(target_month, site_id=None, grade=grade)

    def _get_prior_year_actual_total(self, target_month: str) -> Optional[float]:
        """
        Get actual total volume across all sites and grades for prior year month.

        This queries the database directly for the true historical total,
        ensuring accurate YoY comparisons in BU-level summary reports.

        Args:
            target_month: Target month in 'YYYY-MM' format

        Returns:
            Actual total volume for prior year same month, or None if unavailable
        """
        return self._get_prior_year_actual(target_month, site_id=None, grade=None)

    def _fallback_forecast_value(
        self, data: pd.DataFrame, months_ahead: int, note: Optional[str]
    ) -> float:
        """
        Generate a conservative fallback forecast using median and simple trend.
        """
        recent = data.tail(min(6, len(data)))
        median_recent = float(recent["volume"].median())

        trend = 0.0
        if len(recent) >= 3:
            x = np.arange(len(recent))
            try:
                trend = float(np.polyfit(x, recent["volume"].values, 1)[0])
            except Exception:
                trend = 0.0

        fallback = median_recent + trend * months_ahead
        fallback = max(0.0, fallback)

        logger.info(
            f"Using fallback forecast (median/trend). Months={len(data)}, Note={note or 'none'}"
        )
        return fallback

    def _normalize_forecast_result_types(self, forecast_df: pd.DataFrame) -> pd.DataFrame:
        """
        Normalize optional output columns to stable dtypes.

        This prevents pandas concat warnings when combining many per-site/per-grade
        DataFrames where some optional columns are all-null in a subset of frames.
        """
        df = forecast_df.copy()

        if "snaive_used_fallback" in df.columns:
            df["snaive_used_fallback"] = df["snaive_used_fallback"].astype("boolean")
        if "prior_year_volume" in df.columns:
            df["prior_year_volume"] = pd.to_numeric(
                df["prior_year_volume"], errors="coerce"
            ).astype("Float64")
        if "yoy_change_pct" in df.columns:
            df["yoy_change_pct"] = pd.to_numeric(
                df["yoy_change_pct"], errors="coerce"
            ).astype("Float64")
        if "note" in df.columns:
            df["note"] = df["note"].astype("string")
        for col in ("forecast_p10", "forecast_p90", "interval_width_pct"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").astype("Float64")
        if "risk_flag" in df.columns:
            df["risk_flag"] = df["risk_flag"].astype("string")

        # Sanity-cap columns
        for col in ("forecast_volume_unclamped", "upper_cap_value", "lower_floor_value"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").astype("Float64")
        for col in ("sanity_cap_applied", "yoy_cap_applied"):
            if col in df.columns:
                df[col] = df[col].astype("string")

        return df

    def check_data_sufficiency(
        self, site_id: Optional[str] = None, grade: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Check if there's sufficient data for forecasting

        Returns:
            Dictionary with sufficiency info
        """
        try:
            monthly_data = self.prepare_monthly_data(
                site_id=site_id,
                grade=grade,
                handle_outliers=False,
                fill_gaps=False,
            )
            months_available = len(monthly_data)
            is_sufficient = months_available >= self.min_months_data

            return {
                "sufficient": is_sufficient,
                "months_available": months_available,
                "months_required": self.min_months_data,
                "date_range": f"{monthly_data['date'].min().strftime('%Y-%m')} to {monthly_data['date'].max().strftime('%Y-%m')}",
            }
        except Exception as e:
            return {
                "sufficient": False,
                "months_available": 0,
                "months_required": self.min_months_data,
                "error": str(e),
            }

    def train_models(
        self,
        data: pd.DataFrame,
        models_to_use: Optional[List[str]] = None,
    ) -> Dict[str, object]:
        """
        Train forecasting models

        Args:
            data: Prepared monthly time series data
            models_to_use: List of model names (default: all available)

        Returns:
            Dictionary of successfully trained models
        """
        available_models = get_available_models()

        if models_to_use is None:
            models_to_use = list(available_models.keys())

        trained_models = {}

        for model_name in models_to_use:
            if model_name not in available_models:
                logger.warning(f"Model {model_name} not available")
                continue

            try:
                model = available_models[model_name]()
                model.fit(data)
                trained_models[model_name] = model
                logger.debug(f"Trained model {model_name}")
            except Exception as e:
                logger.warning(f"[x] {model_name}: {e}")

        self.models = trained_models
        return trained_models

    # -- generate_forecast helpers -------------------------------------------

    def _fit_and_predict_model(
        self,
        model_name: str,
        monthly_data: pd.DataFrame,
        monthly_data_raw: pd.DataFrame,
        target_date: pd.Timestamp,
        months_ahead: int,
    ) -> Tuple[float, bool]:
        """Fit *model_name*, predict, and return (forecast_value, snaive_used_fallback).

        Raises on failure so the caller can log and continue.
        """
        available_models = get_available_models()
        model = available_models[model_name]()

        # Use raw data for snaive (preserves actual historical values)
        # Use outlier-handled data for ETS (benefits from cleaned data)
        if model_name == "snaive":
            model.fit(monthly_data_raw)
        else:
            model.fit(monthly_data)

        predictions = model.predict(periods=months_ahead)

        # Get forecast for target month
        target_pred = predictions[predictions["date"] == target_date]
        if target_pred.empty:
            target_pred = predictions.iloc[-1:].copy()

        forecast_value = float(target_pred["forecast"].values[0])
        if not np.isfinite(forecast_value):
            raise ValueError(f"{model_name} produced non-finite forecast")

        # Check if snaive used fallback for the selected target prediction
        used_fallback = False
        if model_name == "snaive":
            if "used_fallback" in target_pred.columns:
                target_used_fallback = target_pred["used_fallback"].iloc[0]
                used_fallback = (
                    bool(target_used_fallback)
                    if pd.notna(target_used_fallback)
                    else False
                )
            elif hasattr(model, "last_prediction_used_fallback"):
                used_fallback = bool(model.last_prediction_used_fallback)

        return forecast_value, used_fallback

    def _build_result_row(
        self,
        model_name: str,
        forecast_value: float,
        target_month: str,
        site_id: Optional[str],
        grade: Optional[str],
        months_available: int,
        data_quality: str,
        note: Optional[str],
        used_fallback: bool,
        show_yoy: bool,
        prior_year_month: Optional[str],
        prior_year_volume: Optional[float],
        site: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Build a single result-row dict, optionally including YoY columns."""
        row: Dict[str, Any] = {
            "model": model_name,
            "target_month": target_month,
            "forecast_volume": forecast_value,
            "site_id": site_id or "ALL",
            "grade": grade or "ALL",
            "months_available": months_available,
            "data_quality": data_quality,
            "note": note,
            "snaive_used_fallback": used_fallback if model_name == "snaive" else None,
        }
        if site is not None:
            row["site"] = site
        if show_yoy:
            row["prior_year_month"] = prior_year_month
            row["prior_year_volume"] = prior_year_volume
            row["yoy_change_pct"] = self._calculate_yoy_change(
                forecast_value, prior_year_volume
            )
        return row

    # -- main entry point ----------------------------------------------------

    def generate_forecast(
        self,
        target_month: str,
        site_id: Optional[str] = None,
        grade: Optional[str] = None,
        models_to_use: Optional[List[str]] = None,
        monthly_data: Optional[pd.DataFrame] = None,
        monthly_data_raw: Optional[pd.DataFrame] = None,
        show_yoy: bool = True,
        prior_year_actuals: Optional[Dict[Tuple[str, str], float]] = None,
        site: Optional[str] = None,
    ) -> pd.DataFrame:
        """
        Generate forecast for a specific month

        Args:
            target_month: Target month in 'YYYY-MM' format
            site_id: Specific site (None for all)
            grade: Specific grade (None for all)
            models_to_use: Specific models to use
            monthly_data: Pre-computed monthly data WITH outlier handling (for ETS)
            monthly_data_raw: Pre-computed monthly data WITHOUT outlier handling (for snaive)
            show_yoy: Include year-over-year comparison columns (default: True)
            prior_year_actuals: Pre-computed lookup from _precompute_prior_year_actuals (bulk mode)

        Returns:
            DataFrame with forecasts from all models
        """
        site_id = _normalize_filter(site_id)
        grade = _normalize_filter(grade)

        # Normalize target month to first day of month
        target_date = pd.to_datetime(target_month).to_period("M").to_timestamp()

        # Get prior year actual for YoY comparison (if enabled)
        prior_year_volume = None
        prior_year_month = None
        if show_yoy:
            if prior_year_actuals is not None:
                key = (str(site_id) if site_id else "ALL", str(grade) if grade else "ALL")
                prior_year_volume = prior_year_actuals.get(key)
            else:
                prior_year_volume = self._get_prior_year_actual(
                    target_month, site_id=site_id, grade=grade
                )
            prior_year_month = (target_date - pd.DateOffset(years=1)).strftime("%Y-%m")

        monthly_data_was_provided = monthly_data is not None
        monthly_data_raw_was_provided = monthly_data_raw is not None

        # Use cached data or prepare fresh
        if monthly_data is None:
            check = self.check_data_sufficiency(site_id=site_id, grade=grade)
            if not check["sufficient"]:
                logger.warning(
                    f"Low data warning: {check['months_available']} months "
                    f"(recommended: {check['months_required']}). Forecasting anyway."
                )
            monthly_data = self.prepare_monthly_data(site_id=site_id, grade=grade, handle_outliers=True)

        if monthly_data_raw is None:
            monthly_data_raw = self.prepare_monthly_data(site_id=site_id, grade=grade, handle_outliers=False, fill_gaps=False)

        last_date = monthly_data["date"].max()
        months_ahead = (target_date.year - last_date.year) * 12 + (
            target_date.month - last_date.month
        )
        if months_ahead <= 0:
            raise ValueError(f"Target month {target_month} is not in the future")

        if monthly_data_raw_was_provided or not monthly_data_was_provided:
            months_available = len(monthly_data_raw)
        else:
            months_available = len(monthly_data)
        data_quality, quality_note = self._assess_data_quality(months_available)

        available_models = get_available_models()
        if models_to_use is None:
            models_to_use = list(available_models.keys())

        # Generate predictions - train each model with appropriate data
        results = []
        forecasts_for_ensemble: Dict[str, float] = {}
        note = quality_note
        fallback_value = None

        for model_name in models_to_use:
            if model_name not in available_models:
                logger.warning(f"Model {model_name} not available")
                continue
            try:
                forecast_value, used_fallback = self._fit_and_predict_model(
                    model_name, monthly_data, monthly_data_raw,
                    target_date, months_ahead,
                )
                forecasts_for_ensemble[model_name] = forecast_value

                model_note = note
                if model_name == "snaive" and used_fallback:
                    fallback_note = (
                        f"SNAIVE used fallback (no same-month data for: "
                        f"{target_date.strftime('%Y-%m')})"
                    )
                    model_note = f"{note}; {fallback_note}" if note else fallback_note
                    logger.debug(fallback_note)

                results.append(self._build_result_row(
                    model_name, forecast_value, target_month,
                    site_id, grade, months_available, data_quality,
                    model_note, used_fallback, show_yoy,
                    prior_year_month, prior_year_volume,
                    site=site,
                ))
            except Exception as e:
                logger.warning(f"Prediction failed for {model_name}: {e}")

        if not forecasts_for_ensemble:
            fallback_value = self._fallback_forecast_value(
                monthly_data, months_ahead, quality_note
            )
            forecasts_for_ensemble["FALLBACK"] = fallback_value
            results.append(self._build_result_row(
                "FALLBACK", fallback_value, target_month,
                site_id, grade, months_available, data_quality,
                note or "Using fallback due to model failure", False,
                show_yoy, prior_year_month, prior_year_volume,
                site=site,
            ))

        results_df = pd.DataFrame(results)

        # Add ensemble row
        ensemble_note = note or ("Used fallback forecast" if fallback_value is not None else None)
        if forecasts_for_ensemble:
            ensemble_forecast = self._compute_ensemble_forecast(
                forecasts_for_ensemble, site_id=site_id
            )
            snaive_fallback_in_results = any(
                r.get("snaive_used_fallback") for r in results if r.get("model") == "snaive"
            )
            ensemble_row_data = self._build_result_row(
                "ENSEMBLE", ensemble_forecast, target_month,
                site_id, grade, months_available, data_quality,
                ensemble_note, False,
                show_yoy, prior_year_month, prior_year_volume,
                site=site,
            )
            # Propagate snaive fallback flag to the ENSEMBLE row so downstream
            # consumers know the ensemble was influenced by a snaive fallback.
            # _build_result_row sets this to None for non-snaive models, so we
            # override it here when snaive actually used fallback.
            ensemble_row_data["snaive_used_fallback"] = (
                True if snaive_fallback_in_results else None
            )

            # Compute prediction intervals for the ensemble
            intervals = self._compute_intervals(
                ensemble_forecast, site_id=site_id, grade=grade,
            )
            ensemble_row_data.update(intervals)

            ensemble_row = pd.DataFrame([ensemble_row_data])
            results_df = pd.concat([results_df, ensemble_row], ignore_index=True)

        return self._normalize_forecast_result_types(results_df)

    # -- post-forecast sanity caps & guardrails --------------------------------

    def _precompute_sanity_bounds(self) -> Dict[tuple, Dict[str, Any]]:
        """Precompute per-(site_id, grade) sanity cap bounds.

        Mature sites (>=MATURE_SITE_MONTHS months) use their own history.
        New sites use grade-cohort stats from mature sites.

        Returns:
            {(site_id, grade): {"upper_cap": float, "lower_floor": float, "is_mature": bool}}
        """
        cohort_stats = self.db.get_grade_cohort_stats()
        site_stats = self.db.get_site_monthly_stats()

        bounds: Dict[tuple, Dict[str, Any]] = {}
        for (site_id, grade), stats in site_stats.items():
            is_mature = stats["months_count"] >= MATURE_SITE_MONTHS
            if is_mature and stats["site_monthly_max"] > 0:
                upper = SANITY_CAP_UPPER_MULT * stats["site_monthly_max"]
                lower = SANITY_CAP_LOWER_MULT * stats["site_monthly_min"] if stats["site_monthly_min"] > 0 else 0.0
            elif grade in cohort_stats and cohort_stats[grade]["cohort_monthly_max"] > 0:
                upper = SANITY_CAP_UPPER_MULT * cohort_stats[grade]["cohort_monthly_max"]
                lower = SANITY_CAP_LOWER_MULT * cohort_stats[grade]["cohort_monthly_min"] if cohort_stats[grade]["cohort_monthly_min"] > 0 else 0.0
            else:
                # No cohort data available — skip capping
                upper = float("inf")
                lower = 0.0

            bounds[(site_id, grade)] = {
                "upper_cap": upper,
                "lower_floor": lower,
                "is_mature": is_mature,
            }

        return bounds

    def _apply_bounds(
        self,
        forecast_value: float,
        site_id: str,
        grade: str,
        bounds: Dict[tuple, Dict[str, Any]],
        monthly_data_raw: Optional[pd.DataFrame],
        prior_year_volume: Optional[float] = None,
    ) -> Tuple[float, float, str, float, float, str]:
        """Clamp *forecast_value* to its sanity bounds and YoY guardrails in one step.

        Sanity bounds come from the site's own history (mature sites) or grade-cohort
        stats (new sites); for new sites with >=3 months of data the upper bound is
        widened via trend extrapolation so legitimately ramping sites aren't neutered.
        YoY guardrails then constrain the result to [LOWER, UPPER] x the prior-year
        actual (skipped when *prior_year_volume* is missing or <= 0).  Applying
        sanity-then-YoY is equivalent to clamping to the intersection of the two
        bounds; YoY wins on the rare conflicting-bound case, matching prior behavior.

        Returns:
            (final_value, unclamped_value, sanity_label, upper_cap_value,
             lower_floor_value, yoy_label)
        """
        unclamped = forecast_value
        info = bounds.get((str(site_id), str(grade)))

        # --- Sanity bounds ---
        if info is None:
            upper_cap, lower_floor = float("inf"), 0.0
        else:
            upper_cap = info["upper_cap"]
            lower_floor = info["lower_floor"]

            # Trend-aware widening for new sites
            if not info["is_mature"] and monthly_data_raw is not None and len(monthly_data_raw) >= 3:
                recent = monthly_data_raw.tail(min(6, len(monthly_data_raw)))
                try:
                    x = np.arange(len(recent))
                    slope = float(np.polyfit(x, recent["volume"].values, 1)[0])
                    if slope > 0:
                        months_ahead = len(monthly_data_raw)  # rough extrapolation distance
                        extrapolated = float(recent["volume"].iloc[-1]) + slope * months_ahead
                        most_recent = float(recent["volume"].iloc[-1])
                        # Cap extrapolation at TREND_EXTRAP_CEILING_MULT * most recent
                        extrapolated = min(extrapolated, TREND_EXTRAP_CEILING_MULT * most_recent)
                        trend_upper = SANITY_CAP_UPPER_MULT * extrapolated
                        if trend_upper > upper_cap:
                            upper_cap = trend_upper
                except Exception:
                    pass

        sanity_label = ""
        if forecast_value > upper_cap:
            forecast_value = upper_cap
            sanity_label = "upper_cap"
        elif forecast_value < lower_floor:
            forecast_value = lower_floor
            sanity_label = "lower_floor"

        # --- YoY guardrails (applied to the sanity-clamped value) ---
        yoy_label = ""
        if prior_year_volume is not None and not pd.isna(prior_year_volume) and prior_year_volume > 0:
            yoy_upper = prior_year_volume * YOY_GUARD_UPPER_RATIO
            yoy_lower = prior_year_volume * YOY_GUARD_LOWER_RATIO
            if forecast_value > yoy_upper:
                forecast_value = yoy_upper
                yoy_label = "yoy_upper"
            elif forecast_value < yoy_lower:
                forecast_value = yoy_lower
                yoy_label = "yoy_lower"

        return forecast_value, unclamped, sanity_label, upper_cap, lower_floor, yoy_label

    def _apply_caps_to_forecast_df(
        self,
        forecast_df: pd.DataFrame,
        sanity_bounds: Dict[tuple, Dict[str, Any]],
        monthly_data_raw: Optional[pd.DataFrame],
    ) -> pd.DataFrame:
        """Apply sanity bounds and YoY guardrails to ENSEMBLE rows in *forecast_df*."""
        df = forecast_df.copy()

        # Initialize new columns with defaults
        df["forecast_volume_unclamped"] = df["forecast_volume"]
        df["sanity_cap_applied"] = ""
        df["upper_cap_value"] = pd.NA
        df["lower_floor_value"] = pd.NA
        df["yoy_cap_applied"] = ""

        ensemble_mask = df["model"] == "ENSEMBLE"
        for idx in df.index[ensemble_mask]:
            row = df.loc[idx]
            prior = row.get("prior_year_volume")
            prior_val = float(prior) if prior is not None and not pd.isna(prior) else None

            vol, unclamped, sanity_label, upper_val, lower_val, yoy_label = self._apply_bounds(
                float(row["forecast_volume"]), str(row["site_id"]), str(row["grade"]),
                sanity_bounds, monthly_data_raw, prior_year_volume=prior_val,
            )

            df.at[idx, "forecast_volume"] = vol
            df.at[idx, "forecast_volume_unclamped"] = unclamped
            df.at[idx, "sanity_cap_applied"] = sanity_label
            df.at[idx, "upper_cap_value"] = upper_val if np.isfinite(upper_val) else pd.NA
            df.at[idx, "lower_floor_value"] = lower_val
            df.at[idx, "yoy_cap_applied"] = yoy_label

            # Recalculate YoY change after a YoY cap fires
            if yoy_label and prior_val and "yoy_change_pct" in df.columns:
                df.at[idx, "yoy_change_pct"] = self._calculate_yoy_change(vol, prior_val)

            # If any cap fired, clip p10/p90 to the effective sanity bounds
            if vol != unclamped:
                for pcol in ("forecast_p10", "forecast_p90"):
                    if pcol in df.columns and pd.notna(df.at[idx, pcol]):
                        pval = float(df.at[idx, pcol])
                        # Clip p10/p90 so they don't exceed the cap
                        if upper_val and np.isfinite(upper_val):
                            pval = min(pval, upper_val)
                        pval = max(pval, lower_val)
                        df.at[idx, pcol] = pval
                # Recalculate interval width and risk flag
                if (
                    "forecast_p10" in df.columns
                    and "forecast_p90" in df.columns
                    and pd.notna(df.at[idx, "forecast_p10"])
                    and pd.notna(df.at[idx, "forecast_p90"])
                    and vol > 0
                ):
                    p10 = float(df.at[idx, "forecast_p10"])
                    p90 = float(df.at[idx, "forecast_p90"])
                    width_pct = (p90 - p10) / vol * 100
                    df.at[idx, "interval_width_pct"] = width_pct
                    if "risk_flag" in df.columns:
                        df.at[idx, "risk_flag"] = (
                            "HIGH" if width_pct > self.RISK_THRESHOLD_PCT else ""
                        )

        return df

    def _create_review_sheet(self, forecasts: pd.DataFrame) -> pd.DataFrame:
        """Create a Review sheet highlighting forecasts that need human attention.

        Includes ENSEMBLE rows where any of: sanity/yoy cap applied,
        sparse/very_sparse data quality, or HIGH risk flag.
        """
        if "model" not in forecasts.columns:
            return pd.DataFrame()

        ensemble = forecasts[forecasts["model"] == "ENSEMBLE"].copy()
        if ensemble.empty:
            return pd.DataFrame()

        # Build mask for rows needing review
        mask = pd.Series(False, index=ensemble.index)

        if "sanity_cap_applied" in ensemble.columns:
            mask |= ensemble["sanity_cap_applied"].fillna("").astype(str).ne("")
        if "yoy_cap_applied" in ensemble.columns:
            mask |= ensemble["yoy_cap_applied"].fillna("").astype(str).ne("")
        if "data_quality" in ensemble.columns:
            mask |= ensemble["data_quality"].isin(("sparse", "very_sparse"))
        if "risk_flag" in ensemble.columns:
            mask |= ensemble["risk_flag"].astype(str).eq("HIGH")

        review = ensemble.loc[mask].copy()
        if review.empty:
            return pd.DataFrame()

        # Sort by severity: capped > very_sparse > sparse > HIGH risk only
        def _severity(row):
            sanity = row.get("sanity_cap_applied", "")
            yoy = row.get("yoy_cap_applied", "")
            # pd.NA or None → treat as empty
            sanity = "" if pd.isna(sanity) else str(sanity)
            yoy = "" if pd.isna(yoy) else str(yoy)
            if sanity != "" or yoy != "":
                return 0
            if row.get("data_quality") == "very_sparse":
                return 1
            if row.get("data_quality") == "sparse":
                return 2
            return 3  # HIGH risk only

        review["_severity"] = review.apply(_severity, axis=1)
        review = review.sort_values("_severity").drop(columns=["_severity"])

        # Select columns for the review sheet
        review_cols = [
            "site_id", "site", "grade", "data_quality", "months_available",
            "forecast_volume", "forecast_volume_unclamped",
            "sanity_cap_applied", "upper_cap_value", "lower_floor_value",
            "yoy_cap_applied", "risk_flag", "interval_width_pct", "note",
        ]
        available = [c for c in review_cols if c in review.columns]
        return review[available].reset_index(drop=True)

    # -- bulk forecasting helpers ---------------------------------------------

    def _forecast_batch(
        self,
        items: pd.DataFrame,
        target_month: str,
        models_to_use: Optional[List[str]],
        show_yoy: bool,
        pya: Optional[Dict[Tuple[str, str], float]],
        skip_insufficient: bool,
        min_months_threshold: int,
        log_every: int,
        label: str,
        sanity_bounds: Optional[Dict[tuple, Dict[str, Any]]] = None,
    ) -> Tuple[List[pd.DataFrame], List[Dict]]:
        """Run generate_forecast for every row in *items* (site/grade combos).

        Each row is expected to have at least 'site_id' and optionally 'grade'
        and 'site'.  Returns (all_forecasts, skipped).
        """
        all_forecasts: List[pd.DataFrame] = []
        skipped: List[Dict] = []
        total = len(items)
        has_grade = "grade" in items.columns

        logger.info(f"Generating forecasts for {total} {label}")

        for i, row in items.iterrows():
            if (i + 1) % log_every == 0 or (i + 1) == total:
                logger.info(f"  Progress: {i+1}/{total} {label}")

            site_id = row.get("site_id")
            grade = row.get("grade") if has_grade else None
            site_name = row.get("site")

            try:
                monthly_data = self.prepare_monthly_data(
                    site_id=site_id, grade=grade, handle_outliers=True,
                )
                monthly_data_raw = self.prepare_monthly_data(
                    site_id=site_id, grade=grade,
                    handle_outliers=False, fill_gaps=False,
                )
                months_available = len(monthly_data_raw)

                if months_available < min_months_threshold and skip_insufficient:
                    skip_info: Dict[str, Any] = {"reason": f"Only {months_available} months"}
                    if site_id:
                        skip_info["site_id"] = site_id
                        skip_info["site"] = row.get("site", "")
                    if grade:
                        skip_info["grade"] = grade
                    skipped.append(skip_info)
                    continue

                forecast = self.generate_forecast(
                    target_month,
                    site_id=site_id,
                    grade=grade,
                    models_to_use=models_to_use,
                    monthly_data=monthly_data,
                    monthly_data_raw=monthly_data_raw,
                    show_yoy=show_yoy,
                    prior_year_actuals=pya,
                    site=site_name,
                )

                # Apply sanity caps and YoY guardrails
                if sanity_bounds is not None:
                    forecast = self._apply_caps_to_forecast_df(
                        forecast, sanity_bounds, monthly_data_raw,
                    )

                all_forecasts.append(forecast)

            except Exception as e:
                parts = []
                if site_id:
                    parts.append(f"Site {site_id}")
                if grade:
                    parts.append(f"Grade {grade}")
                logger.warning(f"  [x] {', '.join(parts) or label}: {e}")
                err_info: Dict[str, Any] = {"reason": str(e)}
                if site_id:
                    err_info["site_id"] = site_id
                    err_info["site"] = row.get("site", "")
                if grade:
                    err_info["grade"] = grade
                skipped.append(err_info)

        return all_forecasts, skipped

    # -- main bulk entry point ------------------------------------------------

    def generate_bulk_forecasts(
        self,
        target_month: str,
        models_to_use: Optional[List[str]] = None,
        output_path: Optional[str] = None,
        skip_insufficient: bool = True,
        show_yoy: bool = True,
    ) -> pd.DataFrame:
        """
        Generate forecasts for multiple configurations with progress tracking

        Args:
            target_month: Target month 'YYYY-MM'
            models_to_use: Specific models to use
            output_path: Output Excel or CSV file path
            skip_insufficient: Skip items with insufficient data
            show_yoy: Include year-over-year comparison columns (default: True)

        Returns:
            DataFrame with all forecasts
        """
        # Precompute all prior-year actuals in one query for the entire run
        pya = self._precompute_prior_year_actuals(target_month) if show_yoy else None

        # Precompute sanity bounds once for the entire run
        sanity_bounds = self._precompute_sanity_bounds()

        items = self.db.get_distinct_site_grades()
        all_forecasts, skipped = self._forecast_batch(
            items, target_month, models_to_use, show_yoy, pya,
            skip_insufficient, self.soft_min_months,
            log_every=50, label="site-grade combinations",
            sanity_bounds=sanity_bounds,
        )

        if not all_forecasts:
            raise ValueError("No forecasts were successfully generated")

        valid_forecasts = [df for df in all_forecasts if not df.empty]
        if not valid_forecasts:
            raise ValueError("No non-empty forecast outputs were generated")
        combined = pd.concat(valid_forecasts, ignore_index=True)

        logger.info("\nForecast Summary:")
        logger.info(f"  [ok] Generated: {len(all_forecasts)} forecasts")
        if skipped:
            logger.info(f"  [skip] Skipped: {len(skipped)} items")

        if output_path:
            self._export_results(combined, skipped, output_path)

        return combined

    def _export_results(
        self, forecasts: pd.DataFrame, skipped: List[Dict], output_path: str
    ):
        """Export forecasts to Excel or CSV based on output extension"""
        suffix = Path(output_path).suffix.lower()
        if suffix == ".csv":
            self._export_to_csv(forecasts, skipped, output_path)
        else:
            self._export_to_excel(forecasts, skipped, output_path)
        logger.info(f"  -> Saved to: {output_path}")

    def _create_site_summary(self, forecasts: pd.DataFrame) -> pd.DataFrame:
        """
        Create site-level summary by summing grade-level forecasts.

        This ensures site totals are reconciled (exactly equal to sum of grades).
        Only applicable when forecasts contain grade-level detail.

        Args:
            forecasts: DataFrame with site-grade detail

        Returns:
            DataFrame with site-level totals derived from summing grades
        """
        # Check if this is grade-level data (has non-ALL grade values)
        if "grade" not in forecasts.columns:
            return pd.DataFrame()

        has_grade_detail = (forecasts["grade"] != "ALL").any()
        if not has_grade_detail:
            return pd.DataFrame()

        # Filter to ENSEMBLE model for clean aggregation (avoid double-counting),
        # consistent with _create_product_summary and _create_bu_summary
        if "model" in forecasts.columns:
            grade_data = forecasts[
                (forecasts["grade"] != "ALL") & (forecasts["model"] == "ENSEMBLE")
            ].copy()
            if grade_data.empty:
                # Fallback to all non-ALL grade data if no ENSEMBLE
                grade_data = forecasts[forecasts["grade"] != "ALL"].copy()
        else:
            grade_data = forecasts[forecasts["grade"] != "ALL"].copy()

        if grade_data.empty:
            return pd.DataFrame()

        # Columns to sum
        sum_cols = ["forecast_volume"]
        if "prior_year_volume" in grade_data.columns:
            sum_cols.append("prior_year_volume")

        # Group by site, sum the volumes across grades
        group_cols = ["site_id", "target_month"]

        # Build aggregation dict
        agg_dict = {col: "sum" for col in sum_cols if col in grade_data.columns}
        agg_dict["grade"] = "count"  # Count of grades included

        site_summary = grade_data.groupby(group_cols, as_index=False).agg(agg_dict)
        site_summary = site_summary.rename(columns={"grade": "grades_included"})

        # Recalculate YoY change based on summed values
        if "prior_year_volume" in site_summary.columns:
            # A summed prior year is misleading when any grade is missing its
            # prior-year actual, so blank the prior year (and YoY) for those sites.
            prior_missing = (
                grade_data.groupby(group_cols)["prior_year_volume"]
                .agg(lambda s: s.isna().any())
                .reset_index(name="_prior_missing")
            )
            site_summary = site_summary.merge(prior_missing, on=group_cols, how="left")
            site_summary.loc[
                site_summary["_prior_missing"].astype(bool), "prior_year_volume"
            ] = np.nan
            site_summary = site_summary.drop(columns="_prior_missing")

            site_summary["yoy_change_pct"] = site_summary.apply(
                lambda row: self._calculate_yoy_change(
                    row["forecast_volume"], row["prior_year_volume"]
                ),
                axis=1
            )

        return _reorder_columns(site_summary, [
            "site_id", "target_month", "forecast_volume",
            "prior_year_volume", "yoy_change_pct", "grades_included",
        ])

    def _create_product_summary(self, forecasts: pd.DataFrame) -> pd.DataFrame:
        """
        Create product/grade-level summary with YoY % change.

        Aggregates forecasts by grade (fuel product type) across all sites,
        showing total volumes and YoY change at the product level.

        IMPORTANT: Prior year volumes are fetched directly from the database
        (not summed from individual forecasts) to ensure accurate YoY comparisons
        even when individual site forecasts used imputed/historical data.

        Args:
            forecasts: DataFrame with forecast data

        Returns:
            DataFrame with product-level totals and YoY %
        """
        if "grade" not in forecasts.columns:
            return pd.DataFrame()

        # Filter to ENSEMBLE model for clean aggregation (avoid double-counting)
        if "model" in forecasts.columns:
            ensemble_data = forecasts[forecasts["model"] == "ENSEMBLE"].copy()
            if ensemble_data.empty:
                # Fallback to all data if no ENSEMBLE
                ensemble_data = forecasts.copy()
        else:
            ensemble_data = forecasts.copy()

        if ensemble_data.empty:
            return pd.DataFrame()

        # Group by grade and target_month, sum forecast volumes
        group_cols = ["grade", "target_month"]
        agg_dict = {"forecast_volume": "sum"}
        if "site_id" in ensemble_data.columns:
            agg_dict["site_id"] = "nunique"  # Count of sites included

        product_summary = ensemble_data.groupby(group_cols, as_index=False).agg(agg_dict)
        if "site_id" in product_summary.columns:
            product_summary = product_summary.rename(columns={"site_id": "sites_included"})

        # Fetch actual prior year volumes directly from database for each grade
        # This ensures accurate YoY comparison regardless of imputation in individual forecasts
        prior_year_volumes = []
        prior_year_months = []
        for _, row in product_summary.iterrows():
            target_month = row["target_month"]
            grade = row["grade"]
            prior_year_vol = self._get_prior_year_actual_by_grade(target_month, grade)
            prior_year_volumes.append(prior_year_vol)
            # Calculate prior year month for display
            target_date = pd.to_datetime(target_month)
            prior_year_months.append(
                (target_date - pd.DateOffset(years=1)).strftime("%Y-%m")
            )

        product_summary["prior_year_month"] = prior_year_months
        product_summary["prior_year_volume"] = prior_year_volumes

        # Calculate YoY change based on actual prior year aggregates
        product_summary["yoy_change_pct"] = product_summary.apply(
            lambda row: self._calculate_yoy_change(
                row["forecast_volume"], row["prior_year_volume"]
            ),
            axis=1
        )

        return _reorder_columns(product_summary, [
            "grade", "target_month", "forecast_volume", "prior_year_month",
            "prior_year_volume", "yoy_change_pct", "sites_included",
        ])

    def _create_bu_summary(self, forecasts: pd.DataFrame) -> pd.DataFrame:
        """
        Create BU (Business Unit) level summary with YoY % change.

        Aggregates all forecasts to show overall totals for the entire BU,
        including total volumes and YoY change.

        IMPORTANT: Prior year volumes are fetched directly from the database
        (not summed from individual forecasts) to ensure accurate YoY comparisons
        even when individual site forecasts used imputed/historical data.

        Args:
            forecasts: DataFrame with forecast data

        Returns:
            DataFrame with BU-level totals and YoY %
        """
        # Filter to ENSEMBLE model for clean aggregation (avoid double-counting)
        if "model" in forecasts.columns:
            ensemble_data = forecasts[forecasts["model"] == "ENSEMBLE"].copy()
            if ensemble_data.empty:
                # Fallback to all data if no ENSEMBLE
                ensemble_data = forecasts.copy()
        else:
            ensemble_data = forecasts.copy()

        if ensemble_data.empty:
            return pd.DataFrame()

        # Group by target_month only (BU-level = all sites, all grades)
        group_cols = ["target_month"]
        agg_dict = {"forecast_volume": "sum"}

        # Count unique sites and grades
        if "site_id" in ensemble_data.columns:
            agg_dict["site_id"] = "nunique"
        if "grade" in ensemble_data.columns:
            agg_dict["grade"] = "nunique"

        bu_summary = ensemble_data.groupby(group_cols, as_index=False).agg(agg_dict)

        # Rename count columns
        if "site_id" in bu_summary.columns:
            bu_summary = bu_summary.rename(columns={"site_id": "sites_included"})
        if "grade" in bu_summary.columns:
            bu_summary = bu_summary.rename(columns={"grade": "grades_included"})

        # Fetch actual prior year total volumes directly from database
        # This ensures accurate YoY comparison regardless of imputation in individual forecasts
        prior_year_volumes = []
        prior_year_months = []
        for _, row in bu_summary.iterrows():
            target_month = row["target_month"]
            prior_year_vol = self._get_prior_year_actual_total(target_month)
            prior_year_volumes.append(prior_year_vol)
            # Calculate prior year month for display
            target_date = pd.to_datetime(target_month)
            prior_year_months.append(
                (target_date - pd.DateOffset(years=1)).strftime("%Y-%m")
            )

        bu_summary["prior_year_month"] = prior_year_months
        bu_summary["prior_year_volume"] = prior_year_volumes

        # Calculate YoY change based on actual prior year total
        bu_summary["yoy_change_pct"] = bu_summary.apply(
            lambda row: self._calculate_yoy_change(
                row["forecast_volume"], row["prior_year_volume"]
            ),
            axis=1
        )

        return _reorder_columns(bu_summary, [
            "target_month", "forecast_volume", "prior_year_month",
            "prior_year_volume", "yoy_change_pct", "sites_included",
            "grades_included",
        ])

    def _load_truck_stop_ids(self) -> List[str]:
        """Load truck stop site IDs from truck_stop_key.xlsx"""
        key_path = Path(__file__).parent / "truck_stop_key.xlsx"
        if not key_path.exists():
            logger.warning(f"Truck stop key not found: {key_path}")
            return []
        truck_df = pd.read_excel(key_path)
        truck_df.columns = truck_df.columns.astype(str).str.strip()
        site_col = next(
            (col for col in truck_df.columns if col.lower() in {"site", "site_id", "siteid"}),
            None,
        )
        if site_col is None:
            logger.warning(f"Truck stop key missing site column: {key_path}")
            return []

        truck_ids = truck_df[site_col].dropna().astype(str).str.strip()
        truck_ids = truck_ids[truck_ids != ""]
        return sorted(truck_ids.unique().tolist())

    def _create_truck_stop_sheet(self, forecasts: pd.DataFrame) -> pd.DataFrame:
        """
        Create Truck Stops sheet: truck stop site IDs with their DSL forecast volume.

        Only populated when forecasts contain grade-level detail.
        """
        has_grade_detail = "grade" in forecasts.columns and (forecasts["grade"] != "ALL").any()
        if not has_grade_detail:
            return pd.DataFrame()

        truck_ids = self._load_truck_stop_ids()
        if not truck_ids:
            return pd.DataFrame()

        required = {"site_id", "model", "grade", "forecast_volume"}
        if not required.issubset(forecasts.columns):
            return pd.DataFrame()

        normalized = forecasts.copy()
        normalized["site_id"] = normalized["site_id"].astype(str).str.strip()

        # Filter to ENSEMBLE DSL forecasts for truck stop sites
        mask = (
            (normalized["model"] == "ENSEMBLE")
            & (normalized["grade"] == "DSL")
            & (normalized["site_id"].isin(truck_ids))
        )
        truck_forecasts = normalized.loc[mask, ["site_id", "forecast_volume"]].copy()
        if not truck_forecasts.empty:
            truck_forecasts = (
                truck_forecasts.groupby("site_id", as_index=False)["forecast_volume"].sum()
            )

        # Ensure all truck stop sites appear (even if no DSL forecast was generated)
        all_truck = pd.DataFrame({"site_id": truck_ids})
        truck_forecasts = all_truck.merge(truck_forecasts, on="site_id", how="left")

        return truck_forecasts

    @staticmethod
    def _format_worksheet(ws):
        """Apply number formatting to a worksheet based on column headers."""
        from openpyxl.utils import get_column_letter

        # Map column names to Excel number formats
        formats = {
            "forecast_volume": "#,##0",
            "forecast_volume_unclamped": "#,##0",
            "upper_cap_value": "#,##0",
            "lower_floor_value": "#,##0",
            "forecast_p10": "#,##0",
            "forecast_p90": "#,##0",
            "prior_year_volume": "#,##0",
            "yoy_change_pct": "0.0",
            "interval_width_pct": "0.0",
            "sites_included": "#,##0",
            "grades_included": "#,##0",
            "months_available": "#,##0",
        }

        # Read header row to find which columns need formatting
        header_map = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value in formats:
                header_map[col_idx] = formats[cell.value]

        # Apply formats to data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column in header_map:
                    cell.number_format = header_map[cell.column]

    def _export_to_excel(
        self, forecasts: pd.DataFrame, skipped: List[Dict], output_path: str
    ):
        """Export forecasts to Excel with multiple sheets"""
        clean_cols = self.CLEAN_COLS
        detail_order = self.DETAIL_COLS

        has_grade_detail = "grade" in forecasts.columns and (forecasts["grade"] != "ALL").any()

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Sheet order: BU Total -> Product Summary -> Site Summary ->
            #              Forecasts -> Truck Stops -> Model Detail -> Skipped

            # BU Total: one-row grand total with YoY
            bu_summary = self._create_bu_summary(forecasts)
            if not bu_summary.empty:
                bu_summary.to_excel(writer, sheet_name="BU Total", index=False)

            # Product Summary & Site Summary: only when grade detail exists
            if has_grade_detail:
                product_summary = self._create_product_summary(forecasts)
                if not product_summary.empty:
                    product_summary.to_excel(writer, sheet_name="Product Summary", index=False)

                site_summary = self._create_site_summary(forecasts)
                if not site_summary.empty:
                    site_summary.to_excel(writer, sheet_name="Site Summary", index=False)

            # Forecasts: ENSEMBLE only, clean columns
            ensemble = forecasts[forecasts["model"] == "ENSEMBLE"].copy()
            if ensemble.empty:
                ensemble = forecasts.copy()
            ensemble_cols = [c for c in clean_cols if c in ensemble.columns]
            ensemble[ensemble_cols].to_excel(
                writer, sheet_name="Forecasts", index=False
            )

            # Truck Stops: DSL forecasts for truck stop sites
            truck_stops = self._create_truck_stop_sheet(forecasts)
            if not truck_stops.empty:
                truck_stops.to_excel(
                    writer, sheet_name="Truck Stops", index=False
                )

            # Review: flagged forecasts needing human attention
            review = self._create_review_sheet(forecasts)
            if not review.empty:
                review.to_excel(writer, sheet_name="Review", index=False)

            # Model Detail: all models with full diagnostic columns
            _reorder_columns(forecasts, list(detail_order)).to_excel(
                writer, sheet_name="Model Detail", index=False
            )

            # Skipped items (if any)
            if skipped:
                skipped_df = pd.DataFrame(skipped)
                skipped_df.to_excel(writer, sheet_name="Skipped", index=False)

            # Apply number formatting to all sheets
            for ws in writer.book.worksheets:
                self._format_worksheet(ws)

    def _export_to_csv(
        self, forecasts: pd.DataFrame, skipped: List[Dict], output_path: str
    ):
        """Export forecasts to CSV; companion files go alongside main file"""
        clean_cols = self.CLEAN_COLS

        ensemble = forecasts[forecasts["model"] == "ENSEMBLE"].copy()
        if ensemble.empty:
            ensemble = forecasts.copy()
        ensemble_cols = [c for c in clean_cols if c in ensemble.columns]
        ensemble[ensemble_cols].to_csv(output_path, index=False)

        base = Path(output_path)
        has_grade_detail = "grade" in forecasts.columns and (forecasts["grade"] != "ALL").any()

        # BU Total
        bu_summary = self._create_bu_summary(forecasts)
        if not bu_summary.empty:
            bu_path = base.with_name(f"{base.stem}_bu_total.csv")
            bu_summary.to_csv(bu_path, index=False)
            logger.info(f"  -> BU Total saved to: {bu_path}")

        # Site Summary and Product Summary: only when grade detail exists
        if has_grade_detail:
            site_summary = self._create_site_summary(forecasts)
            if not site_summary.empty:
                site_summary_path = base.with_name(f"{base.stem}_site_summary.csv")
                site_summary.to_csv(site_summary_path, index=False)
                logger.info(f"  -> Site Summary saved to: {site_summary_path}")

            product_summary = self._create_product_summary(forecasts)
            if not product_summary.empty:
                product_summary_path = base.with_name(f"{base.stem}_product_summary.csv")
                product_summary.to_csv(product_summary_path, index=False)
                logger.info(f"  -> Product Summary saved to: {product_summary_path}")

        # Review: flagged forecasts
        review = self._create_review_sheet(forecasts)
        if not review.empty:
            review_path = base.with_name(f"{base.stem}_review.csv")
            review.to_csv(review_path, index=False)
            logger.info(f"  -> Review saved to: {review_path}")

        # Model Detail: all models with full columns
        detail_path = base.with_name(f"{base.stem}_model_detail.csv")
        _reorder_columns(forecasts, list(self.DETAIL_COLS)).to_csv(detail_path, index=False)
        logger.info(f"  -> Model Detail saved to: {detail_path}")

        if skipped:
            skipped_df = pd.DataFrame(skipped)
            skipped_path = base.with_name(f"{base.stem}_skipped.csv")
            skipped_df.to_csv(skipped_path, index=False)
            logger.info(f"  -> Skipped items saved to: {skipped_path}")

